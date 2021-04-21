# @Time    : 2021/4/19 0019 15:34
# @Author  : 夏末
# @FILE    : python_test.py
# @QQ      : 313098754
# @Company ：个人

import openpyxl   #导入第三方库
# # 函数封装
def read_data(filename,sheetname):      # 定义函数封装
    wb = openpyxl.load_workbook(filename) # 加载工作簿
    sh = wb[sheetname]  # 获取表单
    max_row = sh.max_row  # 获取最大行数
    case_list =[]  # 创建空列表，存储测试用例数据
    for i in range(2,max_row+1):  # 获取所有行数据
        dict1 = dict(
        case_id = sh.cell(row=i,column=1).value,      # 获取case_id
        url = sh.cell(row=i,column=5).value,    # 获取url
        data = sh.cell(row=i,column=6).value,  # 获取data
        expect = sh.cell(row=i,column=7).value  # 获取期望
        )            # 将数据以字典格式输出
        case_list.append(dict1)   # 将每次循环dict数据追加到list中
        print(case_list)
    return case_list   # 定义返回值
# 调用函数
# cases = read_data('test_case_api.xlsx','login')   #调用函数
# print(cases)
# 写入结果
# wb = openpyxl.load_workbook('test_case_api.xlsx')  # 加载工作簿
# sh = wb['register']  # 获取表单
# sh.cell(row=2,column=8).value = 'passed'  # 直接对单元格的内容进行赋值，写入结果
# wb.save('test_case_api.xlsx')   # 将结果保存

# 写入excel测试结果
def wirte_result(filename,sheetname,row,column,final_result):
     wb = openpyxl.load_workbook(filename)  # 加载工作簿
     sh = wb[sheetname]  # 获取表单
     sh.cell(row=row,column=column).value = final_result  # 直接对单元格的内容进行赋值，写入结果
     wb.save(filename)   # 将结果保存